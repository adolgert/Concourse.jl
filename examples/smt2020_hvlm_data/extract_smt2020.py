#!/usr/bin/env python3
"""Extract the SMT2020 dataset-1 (HVLM) workbook into the reduced, minute-
normalized CSVs consumed by ../smt2020_hvlm.jl, and derive per-toolgroup
availability (A = 1 - SDT - UDT).

Self-contained: the only third-party dependency is openpyxl.  No fields contain
commas (verified at the end), so the Julia side parses with a plain split(',').

Usage:
    python3 extract_smt2020.py [path-to-SMT_2020_Model_Data_-_HVLM.xlsx]

If no path is given the default location under examples/SMT2020_data/ is used.
Outputs (written next to this script):
    toolgroups.csv route_product3.csv route_product4.csv lotrelease.csv

All times are normalized to MINUTES.  See README.md for provenance and the
column meaning of each file.  The availability arithmetic (PM + breakdown ->
SDT/UDT/A) is documented in ../../ (analysis.md in the scouting scratchpad) and
reproduces SMT2020 paper Table III to <=0.2 pts on all 11 areas.
"""
import csv
import os
import sys
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_WB = ("/Users/adolgert/dev/Concourse.jl/examples/SMT2020_data/"
              "SMT_2020 - Final/General Data/dataset 1/"
              "SMT_2020_Model_Data_-_HVLM.xlsx")

WAFERS_PER_LOT = 25
WEEK_MIN = 7 * 24 * 60.0                 # 10080 min
REG_INTERVAL = 51.69                     # regular lot of each product (min)
HOT_INTERVAL = 2016.0                    # hot lot of each product (min)
# Offered lots/min per product = regular + hot streams (SuperHotLot excluded;
# see README -- it is a ~0.19% P3 stream and both the paper's 10,000 WSPW target
# and the offered-load oracle use only regular + hot).
LOT_RATE_PER_PRODUCT = 1.0 / REG_INTERVAL + 1.0 / HOT_INTERVAL

TIME_UNITS = {"min": 1.0, "hr": 60.0, "day": 1440.0, "sec": 1.0 / 60.0}


def to_min(value, unit):
    if value is None or value == "":
        return None
    return float(value) * TIME_UNITS[unit]


def blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def read_sheet(wb, name):
    """Return (header, data-rows) stopping at the first all-blank row.  Never
    trust ws.max_row -- dataset 2/4 carry phantom dimensions; stop defensively."""
    ws = wb[name]
    rows, header = [], None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = list(row)
            continue
        if all(blank(c) for c in row):
            break
        rows.append(list(row))
    return header, rows


# ---- Route step schema (column indices in Route_Product_N) --------------------
# 0 ROUTE 1 STEP 2 STEP DESCRIPTION 3 AREA 4 TOOLGROUP 5 PROCESSING UNIT
# 6 PT DIST 7 MEAN 8 OFFSET 9 PT UNITS 10 CASCADING INTERVAL 11 C Units
# 12 BATCH MINIMUM 13 BATCH MAXIMUM 14 SETUP 15 WHEN 16 SETUP DIST 17 SETUP TIME
# 18 OFFSET 19 ST UNITS 20 STEP FOR LTL DEDICATION 21 REWORK PROB % 22 R UNIT
# 23 STEP FOR REWORK 24 PROCESSING PROB % (Sampling) 25 STEP FOR CQT 26 CQT 27 CQT UNITS
class Step:
    __slots__ = ("route", "step", "area", "group", "unit", "pt_min",
                 "pt_off_min", "cascade_min", "batch_min", "batch_max",
                 "rework_prob", "rework_step", "sample_prob")

    def __init__(self, r):
        self.route = r[0]
        self.step = int(r[1])
        self.area = r[3]
        self.group = r[4]
        self.unit = r[5]                                     # Wafer / Lot / Batch
        self.pt_min = to_min(r[7], r[9])
        self.pt_off_min = to_min(r[8], r[9]) if not blank(r[8]) else 0.0
        self.cascade_min = to_min(r[10], r[11]) if not blank(r[10]) else None
        self.batch_min = None if blank(r[12]) else int(r[12])
        self.batch_max = None if blank(r[13]) else int(r[13])
        self.rework_prob = None if blank(r[21]) else float(r[21]) / 100.0
        self.rework_step = None if blank(r[23]) else int(r[23])
        self.sample_prob = 1.0 if blank(r[24]) else float(r[24]) / 100.0


def read_route(wb, name):
    _, rows = read_sheet(wb, name)
    steps = [Step(r) for r in rows]
    assert [s.step for s in steps] == list(range(1, len(steps) + 1)), name
    return steps


def expected_arrivals(steps):
    """Expected arrivals at each step per lot, accounting for rework loops
    (fixed point; probs ~1-2%).  Used only to weight the counter-based PM wafer
    throughput -- rework is NOT modeled in the simulation itself."""
    n = len(steps)
    idx = {s.step: k for k, s in enumerate(steps)}
    fwd = [1.0] * n
    back_targets = [[] for _ in range(n)]
    for k, s in enumerate(steps):
        if s.rework_prob and s.rework_step is not None:
            fwd[k] = 1.0 - s.rework_prob
            back_targets[idx[s.rework_step]].append((k, s.rework_prob))
    f = [0.0] * n
    for _ in range(200):
        newf = [0.0] * n
        newf[0] = 1.0
        for k in range(n):
            if k > 0:
                newf[k] += newf[k - 1] * fwd[k - 1]
            for (src, p) in back_targets[k]:
                newf[k] += f[src] * p
        if max(abs(newf[i] - f[i]) for i in range(n)) < 1e-15:
            f = newf
            break
        f = newf
    return f


def expected_executions(steps):
    arr = expected_arrivals(steps)
    return [arr[k] * steps[k].sample_prob for k in range(len(steps))]


def main():
    wb_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WB
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # ---------------- Toolgroups ----------------
    _, tg_rows = read_sheet(wb, "Toolgroups")
    group_area, group_ntools, group_load, group_unload = {}, {}, {}, {}
    group_cascflag, group_batchflag = {}, {}
    tg_order = []
    for r in tg_rows:
        g = r[1]
        tg_order.append(g)
        group_area[g] = r[0]
        group_ntools[g] = int(r[2])
        group_cascflag[g] = (r[3] == "YES")
        group_batchflag[g] = (r[4] == "YES")
        group_load[g] = to_min(r[7], r[8]) or 0.0
        group_unload[g] = to_min(r[9], r[10]) or 0.0

    # ---------------- Breakdown (per area -> UDT) ----------------
    _, bd_rows = read_sheet(wb, "Breakdown")
    area_udt = {}
    for r in bd_rows:
        mttf = to_min(r[5], r[6])
        mttr = to_min(r[8], r[9])
        area_udt[r[2]] = mttr / (mttf + mttr)

    # ---------------- PM (per toolgroup): SDT ----------------
    _, pm_rows = read_sheet(wb, "PM")
    pm_time_sdt = defaultdict(float)      # summed time-based SDT fraction
    pm_counter = defaultdict(list)        # [(ttr_min, n_wafers), ...]
    for r in pm_rows:
        g = r[2]
        ttr_min = to_min(r[7], r[9])
        if r[3] == "counter based":
            pm_counter[g].append((ttr_min, float(r[4])))     # MTBeforePM in wafers/tool
        else:
            pm_time_sdt[g] += ttr_min / to_min(r[4], r[5])

    # ---------------- Routes ----------------
    routes = {3: read_route(wb, "Route_Product_3"),
              4: read_route(wb, "Route_Product_4")}

    # Offered wafer throughput per tool group (to convert counter-based PMs).
    group_wafer_rate = defaultdict(float)
    for pid, steps in routes.items():
        e = expected_executions(steps)
        for k, s in enumerate(steps):
            group_wafer_rate[s.group] += e[k] * LOT_RATE_PER_PRODUCT * WAFERS_PER_LOT

    # ---------------- Availability per tool group ----------------
    tg_out = []
    for g in tg_order:
        area = group_area[g]
        ntools = group_ntools[g]
        udt = area_udt.get(area, 0.0)
        sdt = pm_time_sdt.get(g, 0.0)
        w_per_tool = group_wafer_rate.get(g, 0.0) / ntools if ntools else 0.0
        for (ttr_min, n_wafers) in pm_counter.get(g, []):
            sdt += ttr_min * w_per_tool / n_wafers
        A = 1.0 - sdt - udt
        tg_out.append([g, area, ntools, f"{A:.6f}",
                       "YES" if group_batchflag[g] else "NO",
                       "YES" if group_cascflag[g] else "NO",
                       f"{group_load[g]:.4f}", f"{group_unload[g]:.4f}"])

    # ---------------- Routes -> reduced CSV ----------------
    def route_rows(steps):
        rows = []
        for s in steps:
            rows.append([
                s.step, s.group, s.unit,
                f"{s.pt_min:.6f}", f"{s.pt_off_min:.6f}",
                "" if s.cascade_min is None else f"{s.cascade_min:.6f}",
                "" if s.batch_min is None else s.batch_min,
                "" if s.batch_max is None else s.batch_max,
                f"{s.sample_prob:.4f}",
            ])
        return rows

    # ---------------- Lotrelease (regular + hot per product) ----------------
    _, lr_rows = read_sheet(wb, "Lotrelease")
    lr_out = []
    for r in lr_rows:
        lot_type = r[2]
        # KEEP only Regular and Hot streams (drop SuperHotLot -- see README).
        if lot_type.startswith("Regular") or lot_type.startswith("HotLot"):
            lr_out.append([r[0], r[1], lot_type, int(r[3]), int(r[5]),
                           f"{to_min(r[8], r[9]):.4f}"])

    # ---------------- Write ----------------
    def write_csv(name, header, rows):
        with open(os.path.join(HERE, name), "w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(header)
            w.writerows(rows)
        # Verify no field contains a comma (plain-split loader on the Julia side).
        for row in [header] + rows:
            for field in row:
                assert "," not in str(field), f"comma in {name}: {field!r}"
        return len(rows)

    n_tg = write_csv("toolgroups.csv",
                     ["toolgroup", "area", "n_tools", "availability",
                      "batching", "cascading", "loading_min", "unloading_min"],
                     tg_out)
    n3 = write_csv("route_product3.csv",
                   ["step", "toolgroup", "unit", "pt_min", "pt_offset_min",
                    "cascade_interval_min", "batch_min_wafers",
                    "batch_max_wafers", "sample_prob"],
                   route_rows(routes[3]))
    n4 = write_csv("route_product4.csv",
                   ["step", "toolgroup", "unit", "pt_min", "pt_offset_min",
                    "cascade_interval_min", "batch_min_wafers",
                    "batch_max_wafers", "sample_prob"],
                   route_rows(routes[4]))
    n_lr = write_csv("lotrelease.csv",
                     ["product", "route", "lot_type", "priority",
                      "wafers_per_lot", "release_interval_min"],
                     lr_out)

    print(f"Wrote (rows excl. header): toolgroups={n_tg} (expect 106) "
          f"route_product3={n3} (expect 583) route_product4={n4} (expect 343) "
          f"lotrelease={n_lr} (expect 4)")
    print("No field contains a comma (asserted).")


if __name__ == "__main__":
    main()
